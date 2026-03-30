
from __future__ import annotations

from pathlib import Path
from datetime import datetime
import openpyxl

import loading_diagram as base
from loading_diagram import EXCEL_FILE


# EXCEL_FILE = "data_(4).xlsx"
SHEET_NAME = "2.a_2)_RIK"


def read_inputs_rik(excel_file: str) -> dict:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb[SHEET_NAME]

    # -----------------------------
    # Table at T5: component masses
    # -----------------------------
    component_name_map = {
        "Wing_new": "Wing",
        "Horizontal Tail": "Horizontal Tail",
        "Vertical Tail": "Vertical Tail",
        "Fuselage_new": "Fuselage",
        "Main Landing gear": "Main Landing gear",
        "Nose Landing gear": "Nose Landing gear",
        "Propulsion System": "Propulsion System",
        "Cockpit Systems": "Cockpit Systems",
        "Batteries front": "Battery front",
        "Battery aft": "Battery aft",
    }

    component_masses: dict[str, float] = {}
    for row in range(6, 16):
        raw_name = ws.cell(row, 20).value  # col T
        if raw_name is None:
            continue
        raw_name = str(raw_name).strip()
        if raw_name in component_name_map:
            component_masses[component_name_map[raw_name]] = base.to_float(
                ws.cell(row, 22).value, f"{raw_name} mass"
            )  # col V

    # -----------------------------
    # Table at Z5: main mass data
    # -----------------------------
    mtow = base.to_float(ws["AB6"].value, "MTOW")
    oew_exc_batt = base.to_float(ws["AB7"].value, "OEW_NEW")
    oew = oew_exc_batt + 4000
    oew_exc_batt_x = base.to_float(ws["W8"].value, "OEW exc batt position from front")
    fuel_mass = base.to_float(ws["AB8"].value, "Fuel weight @ max payload")
    max_payload = base.to_float(ws["AB11"].value, "Max payload")
    pax_total_mass = base.to_float(ws["AB12"].value, "Pax&cabin luggage")
    front_cargo_mass = base.to_float(ws["AB13"].value, "Front cargo hold mass")
    rear_cargo_mass = base.to_float(ws["AB14"].value, "Aft cargo hold mass")

    # -----------------------------
    # Table at T32: lumped loading positions
    # -----------------------------
    fuel_x = base.to_float(ws["V33"].value, "Fuel position from front")
    front_cargo_x = base.to_float(ws["V34"].value, "Front cargo position from front")
    rear_cargo_x = base.to_float(ws["V35"].value, "Rear cargo position from front")
    oew_x_table = base.to_float(ws["X40"].value, "OEW position from front")
    pax_x_table = base.to_float(ws["V39"].value, "PAX position from front")
    mtow_x_table = base.to_float(ws["V40"].value, "MTOW position from front")

    # -----------------------------
    # Table at T42: component CG positions
    # -----------------------------
    component_position_names = {
        "Wing": "Wing",
        "Horizontal Tail": "Horizontal Tail",
        "Vertical Tail": "Vertical Tail",
        "Fuselage": "Fuselage",
        "Main Landing gear": "Main Landing gear",
        "Nose Landing gear": "Nose Landing gear",
        "Propulsion System": "Propulsion System",
        "Cockpit Systems": "Cockpit Systems",
        "Battery front": "Battery front",
        "Battery aft": "Battery aft",
    }

    component_positions: dict[str, float] = {}
    for row in range(43, 53):
        raw_name = ws.cell(row, 20).value  # col T
        if raw_name is None:
            continue
        raw_name = str(raw_name).strip()
        if raw_name in component_position_names:
            component_positions[component_position_names[raw_name]] = base.to_float(
                ws.cell(row, 21).value, f"{raw_name} position from front"
            )  # col U

    nose_to_lemac = base.to_float(ws["U53"].value, "nose to lemac")

    # -----------------------------
    # Table at T57: MAC
    # -----------------------------
    mac = base.to_float(ws["U57"].value, "MAC")

    # -----------------------------
    # Table at T78: passenger geometry
    # -----------------------------
    n_passengers = int(base.to_float(ws["U78"].value, "Number of passengers"))
    n_rows = int(base.to_float(ws["U79"].value, "n rows"))
    seats_per_row = int(base.to_float(ws["U80"].value, "seat per row"))
    seat_pitch = base.to_float(ws["U81"].value, "seat pitch")
    first_row_x = base.to_float(ws["U82"].value, "first row longitudinal position")
    passenger_mass = base.to_float(ws["U83"].value, "passenger mass per person")

    return {
        "mtow": mtow,
        "oew": oew,
        "fuel_mass": fuel_mass,
        "max_payload": max_payload,
        "pax_total_mass_ref": pax_total_mass,
        "front_cargo_mass": front_cargo_mass,
        "rear_cargo_mass": rear_cargo_mass,
        "component_masses": component_masses,
        "component_positions": component_positions,
        "nose_to_lemac": nose_to_lemac,
        "mac": mac,
        "n_passengers": n_passengers,
        "n_rows": n_rows,
        "seats_per_row": seats_per_row,
        "seat_pitch": seat_pitch,
        "first_row_x": first_row_x,
        "passenger_mass": passenger_mass,
        "fuel_x": fuel_x,
        "front_cargo_x": front_cargo_x,
        "rear_cargo_x": rear_cargo_x,
        "oew_x_table": oew_x_table,
        "oew_plot_label": "OEW+batt",
        "oew_plot_xytext": (8, -12),
        "extra_plot_points": [
            {
                "label": "OEW",
                "weight": oew_exc_batt,
                "x_from_front_m": oew_exc_batt_x,
                "xytext": (8, -14),
            }
        ],
        "extra_plot_lines": [
            {
                "from": "OEW",
                "to": "OEW+batt",
                "color": "tab:blue",
                "linewidth": 2.0,
            }
        ],
        "pax_x_table": pax_x_table,
        "mtow_x_table": mtow_x_table,
    }


def main(save_plot: bool = True) -> None:
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / EXCEL_FILE
    if not excel_path.exists():
        raise FileNotFoundError(f"Could not find '{EXCEL_FILE}' next to this script.")

    timestamp = datetime.now()
    timestamp_for_title = timestamp.strftime("%Y-%m-%d %H:%M:%S")
    timestamp_for_filename = timestamp.strftime("%Y-%m-%d_%H-%M-%S")

    output_folder = script_dir / "Loading Diagrams Part 2"
    output_file = output_folder / f"loading_diagram_Part2_{timestamp_for_filename}.png"

    data = read_inputs_rik(str(excel_path))
    paths = base.build_all_paths(data)
    base.print_checks(data, paths)
    base.plot_loading_diagram(
        data=data,
        paths=paths,
        output_folder=output_folder,
        timestamp_for_title=f"{timestamp_for_title} (Part 2)",
        output_file=output_file,
        save_plot=save_plot,
    )

    print(f"\nSaved Part 2 loading diagram to: {output_file}")


if __name__ == "__main__":
    main()
