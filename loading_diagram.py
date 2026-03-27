
from __future__ import annotations

from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
import re
import openpyxl
import matplotlib.pyplot as plt

EXCEL_FILE = "data_(3).xlsx"
REF_SHEET = "1.b) Ref_ac"
MAIN_SHEET = "1.c)"


def to_float(value, name: str) -> float:
    if value is None:
        raise ValueError(f"Missing value for '{name}'.")
    try:
        return float(value)
    except Exception as exc:
        raise ValueError(f"Could not convert '{name}' to float: {value!r}") from exc


def norm(value) -> str:
    return "" if value is None else str(value).strip().lower()


def find_row(ws, label: str, col: int = 2, row_min: int = 1, row_max: int | None = None) -> int:
    row_max = row_max or ws.max_row
    target = norm(label)
    for r in range(row_min, row_max + 1):
        if norm(ws.cell(r, col).value) == target:
            return r
    raise ValueError(f"Could not find '{label}' in sheet '{ws.title}'.")


def x_front_to_percent_mac(x_front: float, nose_to_lemac: float, mac: float) -> float:
    return 100.0 * (x_front - nose_to_lemac) / mac


def cg_update(weight_old: float, x_old: float, weight_item: float, x_item: float) -> tuple[float, float]:
    weight_new = weight_old + weight_item
    x_new = (weight_old * x_old + weight_item * x_item) / weight_new
    return weight_new, x_new


@dataclass
class Point:
    name: str
    weight_kg: float
    x_from_front_m: float
    x_percent_mac: float


def read_inputs(excel_file: str) -> dict:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws_ref = wb[REF_SHEET]
    ws = wb[MAIN_SHEET]

    # 1.b) Ref_ac
    mtow = to_float(ws_ref.cell(find_row(ws_ref, "MTOW", col=2), 4).value, "MTOW")
    oew = to_float(ws_ref.cell(find_row(ws_ref, "OEW", col=2), 4).value, "OEW")
    fuel_mass = to_float(
        ws_ref.cell(find_row(ws_ref, "Fuel weight @ max payload", col=2), 4).value,
        "Fuel weight @ max payload",
    )
    max_payload = to_float(ws_ref.cell(find_row(ws_ref, "Max payload", col=2), 4).value, "Max payload")
    pax_total_mass = to_float(
        ws_ref.cell(find_row(ws_ref, "Pax&cabin luggage", col=3), 4).value,
        "Pax&cabin luggage",
    )
    front_cargo_mass = to_float(
        ws_ref.cell(find_row(ws_ref, "Front cargo hold", col=3), 4).value,
        "Front cargo hold mass",
    )
    rear_cargo_mass = to_float(
        ws_ref.cell(find_row(ws_ref, "Aft cargo hold", col=3), 4).value,
        "Aft cargo hold mass",
    )

    # 1.c) - component masses
    component_names = [
        "Wing",
        "Horizontal Tail",
        "Vertical Tail",
        "Fuselage",
        "Main Landing gear",
        "Nose Landing gear",
        "Propulsion System",
        "Cockpit Systems",
    ]
    component_masses = {}
    for name in component_names:
        r = find_row(ws, name, col=2, row_min=7, row_max=14)
        component_masses[name] = to_float(ws.cell(r, 4).value, f"{name} mass")

    # 1.c) - component CG positions
    component_positions = {}
    for name in component_names:
        r = find_row(ws, name, col=2, row_min=19, row_max=26)
        component_positions[name] = to_float(ws.cell(r, 3).value, f"{name} position from front")

    nose_to_lemac = to_float(ws.cell(find_row(ws, "nose to lemac", col=2, row_min=18, row_max=30), 3).value, "nose to LEMAC")

    # passenger data
    n_passengers = int(to_float(ws.cell(find_row(ws, "number passangers", col=2, row_min=41, row_max=46), 3).value, "number passengers"))
    n_rows = int(to_float(ws.cell(find_row(ws, "n rows", col=2, row_min=41, row_max=46), 3).value, "n rows"))
    seats_per_row = int(to_float(ws.cell(find_row(ws, "seat per row", col=2, row_min=41, row_max=46), 3).value, "seat per row"))
    seat_pitch = to_float(ws.cell(find_row(ws, "seat pitch", col=2, row_min=41, row_max=46), 3).value, "seat pitch")
    first_row_x = to_float(ws.cell(find_row(ws, "first row longitudinal position", col=2, row_min=41, row_max=46), 3).value, "first row longitudinal position")
    passenger_mass = to_float(ws.cell(find_row(ws, "passenger mass per", col=2, row_min=41, row_max=46), 3).value, "passenger mass per")

    # MAC
    mac = ws["F48"].value
    if mac is None:
        text = ws["C49"].value
        if isinstance(text, str):
            match = re.search(r"MAC:\s*([0-9]+(?:\.[0-9]+)?)", text)
            if match:
                mac = float(match.group(1))
    mac = to_float(mac, "MAC")

    # lower loading table rows 54..59
    fuel_row = find_row(ws, "Fuel", col=2, row_min=54, row_max=59)
    front_cargo_row = find_row(ws, "Front cargo compart", col=2, row_min=54, row_max=59)
    rear_cargo_row = find_row(ws, "Rear Crago Comp", col=2, row_min=54, row_max=59)
    oew_row = find_row(ws, "OEW", col=2, row_min=54, row_max=59)
    pax_row = find_row(ws, "PAX", col=2, row_min=54, row_max=59)
    mtow_row = find_row(ws, "MTOW", col=2, row_min=54, row_max=59)

    fuel_x = to_float(ws.cell(fuel_row, 4).value, "Fuel position from front")
    front_cargo_x = to_float(ws.cell(front_cargo_row, 4).value, "Front cargo position from front")
    rear_cargo_x = to_float(ws.cell(rear_cargo_row, 4).value, "Rear cargo position from front")
    oew_x_table = to_float(ws.cell(oew_row, 4).value, "OEW position from front")
    pax_x_table = to_float(ws.cell(pax_row, 4).value, "PAX position from front")
    mtow_x_table = to_float(ws.cell(mtow_row, 4).value, "MTOW position from front")

    return {
        "mtow": mtow,
        "oew": oew,
        "fuel_mass": fuel_mass,
        "max_payload": max_payload,
        "pax_total_mass_ref": pax_total_mass,
        "front_cargo_mass": front_cargo_mass,
        "rear_cargo_mass": rear_cargo_mass,
        "component_masses": component_masses,
        "component_positions": component_positions,
        "nose_to_lemac": nose_to_lemac,
        "mac": mac,
        "n_passengers": n_passengers,
        "n_rows": n_rows,
        "seats_per_row": seats_per_row,
        "seat_pitch": seat_pitch,
        "first_row_x": first_row_x,
        "passenger_mass": passenger_mass,
        "fuel_x": fuel_x,
        "front_cargo_x": front_cargo_x,
        "rear_cargo_x": rear_cargo_x,
        "oew_x_table": oew_x_table,
        "pax_x_table": pax_x_table,
        "mtow_x_table": mtow_x_table,
    }


def make_point(name: str, weight: float, x_front: float, data: dict) -> Point:
    return Point(
        name=name,
        weight_kg=weight,
        x_from_front_m=x_front,
        x_percent_mac=x_front_to_percent_mac(x_front, data["nose_to_lemac"], data["mac"]),
    )


def compute_oew_cg(component_masses: dict[str, float], component_positions: dict[str, float]) -> float:
    total_mass = sum(component_masses.values())
    total_moment = sum(component_masses[k] * component_positions[k] for k in component_masses)
    return total_moment / total_mass


def row_positions(first_row_x: float, n_rows: int, seat_pitch: float) -> list[float]:
    return [first_row_x + i * seat_pitch for i in range(n_rows)]


def extend_with_item(path: list[Point], item_name: str, item_mass: float, item_x: float, data: dict) -> list[Point]:
    last = path[-1]
    w_new, x_new = cg_update(last.weight_kg, last.x_from_front_m, item_mass, item_x)
    return path + [make_point(item_name, w_new, x_new, data)]


def build_cargo_paths(data: dict, oew_x: float) -> list[tuple[str, list[Point]]]:
    start = [make_point("OEW", data["oew"], oew_x, data)]

    p1 = extend_with_item(start, "Front cargo", data["front_cargo_mass"], data["front_cargo_x"], data)
    p1 = extend_with_item(p1, "Rear cargo", data["rear_cargo_mass"], data["rear_cargo_x"], data)

    p2 = extend_with_item(start, "Rear cargo", data["rear_cargo_mass"], data["rear_cargo_x"], data)
    p2 = extend_with_item(p2, "Front cargo", data["front_cargo_mass"], data["front_cargo_x"], data)

    return [
        ("Cargo: front→rear", p1),
        ("Cargo: rear→front", p2),
    ]


def build_passenger_branch(base_path: list[Point], label_prefix: str, positions: list[float], mass_per_row: float, data: dict) -> list[Point]:
    path = list(base_path)
    for i, x in enumerate(positions, start=1):
        path = extend_with_item(path, f"{label_prefix} row {i}", mass_per_row, x, data)
    return path


def build_all_paths(data: dict) -> dict:
    oew_x = compute_oew_cg(data["component_masses"], data["component_positions"])
    seat_xs = row_positions(data["first_row_x"], data["n_rows"], data["seat_pitch"])

    # Use row-based loading mass exactly as modeled in the cabin
    window_mass_per_row = 2 * data["passenger_mass"]
    aisle_mass_per_row = 2 * data["passenger_mass"]

    cargo_paths = build_cargo_paths(data, oew_x)
    window_paths = []
    aisle_paths = []
    fuel_paths = []

    for cargo_name, cargo_path in cargo_paths:
        w_fb = build_passenger_branch(cargo_path, "Window", seat_xs, window_mass_per_row, data)
        w_bf = build_passenger_branch(cargo_path, "Window", list(reversed(seat_xs)), window_mass_per_row, data)

        window_paths.append((f"{cargo_name} + windows front→back", w_fb))
        window_paths.append((f"{cargo_name} + windows back→front", w_bf))

        for w_name, w_path in [
            (f"{cargo_name} + windows front→back", w_fb),
            (f"{cargo_name} + windows back→front", w_bf),
        ]:
            a_fb = build_passenger_branch(w_path, "Aisle", seat_xs, aisle_mass_per_row, data)
            a_bf = build_passenger_branch(w_path, "Aisle", list(reversed(seat_xs)), aisle_mass_per_row, data)

            aisle_paths.append((f"{w_name} + aisles front→back", a_fb))
            aisle_paths.append((f"{w_name} + aisles back→front", a_bf))

            f1 = extend_with_item(a_fb, "Fuel", data["fuel_mass"], data["fuel_x"], data)
            f2 = extend_with_item(a_bf, "Fuel", data["fuel_mass"], data["fuel_x"], data)

            fuel_paths.append((f"{w_name} + aisles front→back + fuel", f1))
            fuel_paths.append((f"{w_name} + aisles back→front + fuel", f2))

    return {
        "oew_x": oew_x,
        "cargo_paths": cargo_paths,
        "window_paths": window_paths,
        "aisle_paths": aisle_paths,
        "fuel_paths": fuel_paths,
    }


def all_points(paths_group: list[tuple[str, list[Point]]]) -> list[tuple[str, Point]]:
    pts = []
    for name, path in paths_group:
        for p in path:
            pts.append((name, p))
    return pts


def get_extremes(all_path_groups: dict) -> dict:
    pts = []
    for key in ["cargo_paths", "window_paths", "aisle_paths", "fuel_paths"]:
        pts.extend(all_points(all_path_groups[key]))

    most_forward = min(pts, key=lambda item: item[1].x_percent_mac)
    most_aft = max(pts, key=lambda item: item[1].x_percent_mac)
    return {"most_forward": most_forward, "most_aft": most_aft}


def print_checks(data: dict, paths: dict) -> None:
    oew_x = paths["oew_x"]
    print("\n========== SANITY CHECKS ==========")
    print(f"OEW CG from component build-up : {oew_x:.4f} m from nose")
    print(f"OEW CG from lower table        : {data['oew_x_table']:.4f} m from nose")
    print(f"Difference                     : {oew_x - data['oew_x_table']:+.6f} m")

    modeled_pax = data["n_passengers"] * data["passenger_mass"]
    print(f"\nPassenger mass from row model   : {modeled_pax:.1f} kg")
    print(f"Passenger mass from ref sheet   : {data['pax_total_mass_ref']:.1f} kg")
    print(f"Difference                      : {modeled_pax - data['pax_total_mass_ref']:+.1f} kg")

    print("\nKnown loading positions:")
    for label, x in [
        ("Front cargo", data["front_cargo_x"]),
        ("Rear cargo", data["rear_cargo_x"]),
        ("Fuel", data["fuel_x"]),
        ("OEW", oew_x),
        ("PAX(table)", data["pax_x_table"]),
        ("MTOW(table)", data["mtow_x_table"]),
        ("First row", data["first_row_x"]),
    ]:
        print(f"{label:12s}: {x:8.4f} m from nose   {x_front_to_percent_mac(x, data['nose_to_lemac'], data['mac']):8.3f} %MAC")

    final_weights = [path[-1].weight_kg for _, path in paths["fuel_paths"]]
    print(f"\nFinal loaded weight (all branches): {min(final_weights):.1f} to {max(final_weights):.1f} kg")
    if max(final_weights) > data["mtow"]:
        print(f"WARNING: exceeds MTOW by {max(final_weights) - data['mtow']:.1f} kg")

    ext = get_extremes(paths)
    f_name, f_point = ext["most_forward"]
    a_name, a_point = ext["most_aft"]

    print("\n========== EXTREMES ==========")
    print(f"Most forward CG : {f_point.x_percent_mac:.3f} %MAC at {f_point.weight_kg:.1f} kg")
    print(f"  path          : {f_name}")
    print(f"  point         : {f_point.name}")

    print(f"Most aft CG     : {a_point.x_percent_mac:.3f} %MAC at {a_point.weight_kg:.1f} kg")
    print(f"  path          : {a_name}")
    print(f"  point         : {a_point.name}")


def plot_loading_diagram(
    data: dict,
    paths: dict,
    output_folder: Path,
    timestamp_for_title: str,
    output_file: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(12, 8))

    # -------- Cargo: blue --------
    cargo_label_added = False
    for name, path in paths["cargo_paths"]:
        x = [p.x_percent_mac for p in path]
        y = [p.weight_kg for p in path]
        ax.plot(
            x, y,
            color="tab:blue",
            marker="o",
            linewidth=2.0,
            markersize=4,
            label="Cargo" if not cargo_label_added else None
        )
        cargo_label_added = True

    # -------- Window seats: orange --------
    window_label_added = False
    for name, path in paths["window_paths"]:
        x = [p.x_percent_mac for p in path[2:]]
        y = [p.weight_kg for p in path[2:]]
        ax.plot(
            x, y,
            color="tab:orange",
            marker="x",
            linewidth=1.8,
            markersize=4,
            alpha=0.95,
            label="Window seats" if not window_label_added else None
        )
        window_label_added = True

    # -------- Aisle seats: green --------
    aisle_label_added = False
    for name, path in paths["aisle_paths"]:
        x = [p.x_percent_mac for p in path[-(data["n_rows"] + 1):]]
        y = [p.weight_kg for p in path[-(data["n_rows"] + 1):]]
        ax.plot(
            x, y,
            color="tab:green",
            marker="s",
            linewidth=1.6,
            markersize=3.5,
            alpha=0.9,
            label="Aisle seats" if not aisle_label_added else None
        )
        aisle_label_added = True

    # -------- Fuel: red --------
    fuel_label_added = False
    for name, path in paths["fuel_paths"]:
        x = [p.x_percent_mac for p in path[-2:]]
        y = [p.weight_kg for p in path[-2:]]
        ax.plot(
            x, y,
            color="tab:red",
            marker="D",
            linewidth=2.0,
            markersize=5,
            alpha=0.95,
            label="Fuel" if not fuel_label_added else None
        )
        fuel_label_added = True

    # -------- Key markers --------
    oew = make_point("OEW", data["oew"], paths["oew_x"], data)
    ax.scatter([oew.x_percent_mac], [oew.weight_kg], color="black", s=70, zorder=5)
    ax.annotate("OEW", (oew.x_percent_mac, oew.weight_kg), xytext=(8, 8), textcoords="offset points")

    ext = get_extremes(paths)
    _, p_fwd = ext["most_forward"]
    _, p_aft = ext["most_aft"]

    ax.scatter([p_fwd.x_percent_mac], [p_fwd.weight_kg], color="red", s=75, zorder=6)
    ax.annotate("Most forward", (p_fwd.x_percent_mac, p_fwd.weight_kg), xytext=(8, -14), textcoords="offset points")

    ax.scatter([p_aft.x_percent_mac], [p_aft.weight_kg], color="red", s=75, zorder=6)
    ax.annotate("Most aft", (p_aft.x_percent_mac, p_aft.weight_kg), xytext=(8, 8), textcoords="offset points")

    ax.set_xlabel("x_cg [%MAC]")
    ax.set_ylabel("Mass [kg]")
    ax.set_title(f"Loading Diagram - {timestamp_for_title}")
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()

    output_folder.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_file, dpi=300)
    plt.show()


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / EXCEL_FILE
    if not excel_path.exists():
        raise FileNotFoundError(f"Could not find '{EXCEL_FILE}' next to this script.")

    timestamp = datetime.now()
    timestamp_for_title = timestamp.strftime("%Y-%m-%d %H:%M:%S")
    timestamp_for_filename = timestamp.strftime("%Y-%m-%d_%H-%M-%S")

    output_folder = script_dir / "Loading Diagrams"
    output_file = output_folder / f"loading_diagram_{timestamp_for_filename}.png"

    data = read_inputs(str(excel_path))
    paths = build_all_paths(data)
    print_checks(data, paths)
    plot_loading_diagram(
        data=data,
        paths=paths,
        output_folder=output_folder,
        timestamp_for_title=timestamp_for_title,
        output_file=output_file,
    )

    print(f"\nSaved loading diagram to: {output_file}")


if __name__ == "__main__":
    main()
